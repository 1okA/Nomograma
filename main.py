from tkinter import *
from tkinter import filedialog
import matplotlib
import win32com.client as win32
from openpyxl import load_workbook

# Обрезваем переменную до установленного придела
def toFixed(numObj, digits=0):
    return f"{numObj:.{digits}f}"

#Делаем загрузку эксель из выбора файла(светофоры)
#Проверяем не пустая ли строка
def is_not_blank(s):
    return bool(s and s.strip())

# Удаляем лишний хлам из строки
def cleanstring(str):
    import re
    str = re.sub("^\s+|\n|\r|\s+$", '', str)
    return str

def clicked():
    # Получаем доступ к файлу Excel
    fn = filedialog.askopenfilename()
    wb = load_workbook(fn)
    ws = wb["Лист1"]
    listR = []

    for i in range(2, ws.max_row):
        ords = ws.cell(row=i, column=3).value
        listR.append({
            # Получаем регион
            'station': ws.cell(row=i, column=1).value,
            'parameter': ws.cell(row=i, column=2).value,
            'ordinate': int(ords.split('+', 1)[0]) + int(ords.split('+', 1)[1]) / 1000
        })

    print(listR)
    clicked2(listR)

#Вывод графика
def clicked2(listR):
    print(listR)
    visio = win32.gencache.EnsureDispatch('Visio.Application')
    document = visio.Documents.Add("")
    #
    page = visio.ActivePage
    #
    shapes = page.Shapes

    # Отрисовка блокоф сфетофоров
    y = 0
    x = 0
    y1 = 1
    x1 = 2 / 5
    # Отрисовка блоков координат (киллометры)
    y2 = 6
    x2 = 0
    y21 = 7
    x21 = 2/5
    p = 0

    # Количество ячеек общее
    g = len(listR)
    print(g)

    for i in range(g):
        # Выводим объект 1
        rect1 = page.DrawRectangle(x, y, x1, y1)
        rect1.Text = listR[p]['parameter']
        # Выводим объект 2
        rect2 = page.DrawRectangle(x2, y2, x21, y21)
        rect2.Text = toFixed(listR[p]['ordinate'])

        x2 = x2+2/5
        x21 = x21+2/5
        p = p+1
        x = x+2/5
        x1 = x1+2/5




#Отрисовываем модель окна
window = Tk()
window.title("Добро пожаловать!")
window.geometry('600x400')
matplotlib.use('TkAgg')
btn = Button(window, text="Выбрать Файл", command=clicked)
btn.grid(column=2, row=0)
btn = Button(window, text="Visio", command=clicked2)
btn.grid(column=4, row=0)
matplotlib.use('TkAgg')
window.mainloop()





